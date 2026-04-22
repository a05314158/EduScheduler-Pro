"""export_excel.py — экспорт расписания в schedule.xlsx (три листа)."""
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

from slots import DAY_NAMES_SHORT

# ── Цвета ──────────────────────────────────────────────────────────────────
CLR_HEADER_BG  = "2E4057"   # тёмно-синий заголовок
CLR_HEADER_FG  = "FFFFFF"   # белый текст заголовка
CLR_ROW_ODD    = "F2F5FA"   # светло-голубой нечётные строки
CLR_ROW_EVEN   = "FFFFFF"   # белый чётные строки
CLR_DAY_LABEL  = "D6E4F0"   # голубой метка дня (первый столбец)
CLR_SUMMARY_H  = "1B4F72"   # тёмный заголовок сводки
CLR_ACCENT     = "2471A3"   # акцент

THIN = Side(style="thin", color="CCCCCC")
BORDER_CELL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hfill(color):
    return PatternFill("solid", fgColor=color)


def _header_style(cell, text, color=CLR_HEADER_BG):
    cell.value = text
    cell.font  = Font(bold=True, color=CLR_HEADER_FG, size=10)
    cell.fill  = _hfill(color)
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_CELL


def _data_style(cell, text, row_idx, align=ALIGN_LEFT):
    cell.value = text
    bg = CLR_ROW_ODD if row_idx % 2 == 1 else CLR_ROW_EVEN
    cell.fill = _hfill(bg)
    cell.alignment = align
    cell.border = BORDER_CELL
    cell.font = Font(size=9)


def _auto_width(ws, min_w=10, max_w=40):
    """Авторазмер столбцов по содержимому."""
    for col in ws.columns:
        length = min_w
        for cell in col:
            if cell.value:
                # многострочное значение — берём максимальную строку
                lines = str(cell.value).split("\n")
                length = max(length, max(len(l) for l in lines) + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length, max_w)


def _build_slot_labels(slots):
    """Список уникальных меток слотов в порядке (day, period)."""
    seen, labels = set(), []
    for slot in sorted(slots, key=lambda s: (s.day, s.period)):
        key = (slot.day, slot.period)
        if key not in seen:
            seen.add(key)
            labels.append((slot.day, slot.period,
                           f"{DAY_NAMES_SHORT[slot.day]} п{slot.period+1}",
                           slot.time_range))
    return labels


# ── Лист 1: По группам ─────────────────────────────────────────────────────
def _sheet_groups(wb, assigned, lessons, slots, groups):
    ws = wb.create_sheet("По группам")
    ws.freeze_panes = "B3"

    slot_labels = _build_slot_labels(slots)
    sorted_groups = sorted(groups, key=lambda g: g.name)

    # Строка 1 — заголовки групп
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    _header_style(ws.cell(1, 1), "День / Пара")
    _header_style(ws.cell(2, 1), "Время")
    for col_idx, group in enumerate(sorted_groups, start=2):
        _header_style(ws.cell(1, col_idx), group.name)
        _header_style(ws.cell(2, col_idx), f"смена {group.shift}")

    # Индекс: (group_name, day, period) -> "Предмет\nПреподаватель"
    cell_data = defaultdict(str)
    for (l_idx, s_idx), _ in assigned.items():
        les  = lessons[l_idx]
        slot = slots[s_idx]
        text = f"{les.subject.name}\n{les.teacher.name}"
        cell_data[(les.group.name, slot.day, slot.period)] = text

    for row_idx, (day, period, label, trange) in enumerate(slot_labels, start=3):
        ws.row_dimensions[row_idx].height = 32

        # Метка дня/пары
        lbl_cell = ws.cell(row_idx, 1, label)
        lbl_cell.font   = Font(bold=True, size=9)
        lbl_cell.fill   = _hfill(CLR_DAY_LABEL)
        lbl_cell.alignment = ALIGN_CENTER
        lbl_cell.border = BORDER_CELL

        # Время
        # (уже вписано в label, ставим во второй строке — но у нас нет отдельной колонки времени)
        # Добавим время в подпись ячейки первого столбца
        lbl_cell.value  = f"{label}\n{trange}"

        for col_idx, group in enumerate(sorted_groups, start=2):
            text = cell_data.get((group.name, day, period), "")
            _data_style(ws.cell(row_idx, col_idx), text, row_idx - 2)

    _auto_width(ws, min_w=12, max_w=28)
    ws.column_dimensions["A"].width = 16


# ── Лист 2: По преподавателям ──────────────────────────────────────────────
def _sheet_teachers(wb, assigned, lessons, slots, teachers):
    ws = wb.create_sheet("По преподавателям")
    ws.freeze_panes = "B2"

    slot_labels = _build_slot_labels(slots)
    sorted_teachers = sorted(teachers, key=lambda t: t.name)

    # Заголовок
    ws.row_dimensions[1].height = 30
    _header_style(ws.cell(1, 1), "День / Пара / Время")
    for col_idx, teacher in enumerate(sorted_teachers, start=2):
        # Короткое имя: "Иванов И. И." → оставляем как есть (уже краткое)
        _header_style(ws.cell(1, col_idx), teacher.name)

    # Индекс: (teacher_name, day, period) -> "Группа\nПредмет"
    cell_data = defaultdict(str)
    for (l_idx, s_idx), _ in assigned.items():
        les  = lessons[l_idx]
        slot = slots[s_idx]
        text = f"{les.group.name}\n{les.subject.name}"
        cell_data[(les.teacher.name, slot.day, slot.period)] = text

    for row_idx, (day, period, label, trange) in enumerate(slot_labels, start=2):
        ws.row_dimensions[row_idx].height = 32
        lbl_cell = ws.cell(row_idx, 1, f"{label}\n{trange}")
        lbl_cell.font      = Font(bold=True, size=9)
        lbl_cell.fill      = _hfill(CLR_DAY_LABEL)
        lbl_cell.alignment = ALIGN_CENTER
        lbl_cell.border    = BORDER_CELL

        for col_idx, teacher in enumerate(sorted_teachers, start=2):
            text = cell_data.get((teacher.name, day, period), "")
            _data_style(ws.cell(row_idx, col_idx), text, row_idx - 1)

    _auto_width(ws, min_w=14, max_w=26)
    ws.column_dimensions["A"].width = 16


# ── Лист 3: Сводка ─────────────────────────────────────────────────────────
def _sheet_summary(wb, assigned, lessons, slots, groups):
    ws = wb.create_sheet("Сводка")
    ws.freeze_panes = "A2"

    headers = ["Группа", "Смена", "Пар в неделю", "Дней с занятиями", "Дни занятий"]
    ws.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers, start=1):
        _header_style(ws.cell(1, col_idx), h, color=CLR_SUMMARY_H)

    # Собираем данные по группам
    group_stats = defaultdict(lambda: {"pairs": 0, "days": set()})
    for (l_idx, s_idx), _ in assigned.items():
        les  = lessons[l_idx]
        slot = slots[s_idx]
        group_stats[les.group.name]["pairs"] += 1
        group_stats[les.group.name]["days"].add(slot.day)

    sorted_groups = sorted(groups, key=lambda g: g.name)
    for row_idx, group in enumerate(sorted_groups, start=2):
        ws.row_dimensions[row_idx].height = 18
        stats = group_stats.get(group.name, {"pairs": 0, "days": set()})
        days_str = ", ".join(DAY_NAMES_SHORT[d] for d in sorted(stats["days"]))
        n_days   = len(stats["days"])

        row_data = [group.name, group.shift, stats["pairs"], n_days, days_str]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            bg = CLR_ROW_ODD if (row_idx % 2 == 1) else CLR_ROW_EVEN
            cell.fill = _hfill(bg)
            cell.alignment = ALIGN_LEFT if col_idx in (1, 5) else ALIGN_CENTER
            cell.border = BORDER_CELL
            cell.font = Font(size=10,
                             bold=(col_idx == 1),
                             color=("000000"))

    col_widths = [18, 8, 14, 18, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Публичный API ──────────────────────────────────────────────────────────
def export_excel(solver, assignments, lessons, slots, groups, teachers,
                 output_path="schedule.xlsx"):
    """
    Генерирует schedule.xlsx с тремя листами.

    Вызывается после solve() и assign_rooms(), принимает те же объекты.
    """
    # Собираем assigned (только назначенные пары)
    assigned = {
        (l_idx, s_idx): True
        for (l_idx, s_idx), var in assignments.items()
        if solver.value(var) == 1
    }

    wb = Workbook()
    wb.remove(wb.active)          # удаляем дефолтный пустой лист

    _sheet_groups(wb, assigned, lessons, slots, groups)
    _sheet_teachers(wb, assigned, lessons, slots, teachers)
    _sheet_summary(wb, assigned, lessons, slots, groups)

    wb.save(output_path)
    print(f"[EXCEL] Сохранено: {output_path}  "
          f"(листы: По группам / По преподавателям / Сводка)")
    return output_path
